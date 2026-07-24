"""
Excel export of the HAZOP preparation worksheet — in the format a HAZOP
meeting actually uses, so the deliverable is recognisable as "our worksheet,
pre-filled" rather than a developer CSV.

Layout per workbook:
  "Summary"       one row per node: members, row count, rows with a found
                  safeguard, review status counts
  one sheet/node  the worksheet rows for that node with the standard HAZOP
                  columns (Deviation | Causes | Consequences | Safeguards |
                  Recommendation | Action party | Status), frozen header,
                  wrapped text, sensible column widths, status colouring

Only openpyxl is used (no pandas dependency here), and the function takes
the same row dicts build_worksheet produces — including any edits made in
the Streamlit data editor, which is the point: export follows review.
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_COLS = [("deviation", "Deviation", 18), ("causes", "Causes", 52),
         ("consequences", "Consequences", 52), ("safeguards", "Safeguards", 28),
         ("severity", "S", 5), ("likelihood", "L", 5), ("risk", "Risk", 12),
         ("recommendation", "Recommendation", 34),
         ("action_party", "Action party", 16), ("status", "Status", 12)]

_HEAD_FILL = PatternFill("solid", fgColor="1F3864")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_STATUS_FILL = {"proposed": PatternFill("solid", fgColor="FFF2CC"),
                "reviewed": PatternFill("solid", fgColor="E2EFDA"),
                "rejected": PatternFill("solid", fgColor="F8CBAD")}
_WRAP = Alignment(wrap_text=True, vertical="top")


def _sheet_name(node: str, used: set[str]) -> str:
    """Excel sheet names: <=31 chars, no []:*?/\\ — and unique."""
    base = re.sub(r"[\[\]:*?/\\]", "-", node)[:31] or "node"
    name, i = base, 2
    while name in used:
        suffix = f"~{i}"
        name, i = base[:31 - len(suffix)] + suffix, i + 1
    used.add(name)
    return name


def write_worksheet_xlsx(rows: list[dict], path: Path,
                         title: str = "HAZOP preparation",
                         meta: dict | None = None) -> Path:
    """One formatted sheet per node + a summary sheet. Rows may carry edits
    (recommendation / action_party / status) from the review UI."""
    by_node: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_node[r["node"]].append(r)

    wb = Workbook()
    used: set[str] = set()

    # ---- summary sheet -----------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    if meta and any(str(v).strip() for v in meta.values()):
        for k, lab in (("chairman", "Chairman"), ("date", "Meeting date"),
                       ("participants", "Participants"),
                       ("revision", "Revision")):
            if str(meta.get(k, "")).strip():
                ws.append([f"{lab}: {meta[k]}"])
                ws.cell(row=ws.max_row, column=1).font = Font(size=10)
    ws.append([])
    head = ["Node", "Members", "Rows", "Rows w/ safeguard",
            "proposed", "reviewed", "rejected"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT
    for node, rs in sorted(by_node.items()):
        sg = sum(1 for r in rs if not r["safeguards"].startswith("(none"))
        st = {k: sum(1 for r in rs if r.get("status", "proposed") == k)
              for k in ("proposed", "reviewed", "rejected")}
        ws.append([node, rs[0].get("node_members", ""), len(rs), sg,
                   st["proposed"], st["reviewed"], st["rejected"]])
    for col, w in zip("ABCDEFG", (26, 60, 8, 18, 10, 10, 10)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=4):
        row[1].alignment = _WRAP
    ws.freeze_panes = "A4"
    ws.append([])
    ws.append(["AI-prepared draft from extracted P&ID/SCD data — "
               "for HAZOP team review, not a completed study."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)

    # ---- one sheet per node ------------------------------------------------
    for node, rs in sorted(by_node.items()):
        s = wb.create_sheet(_sheet_name(node, used))
        s.append([f"Node: {node}"])
        s["A1"].font = Font(bold=True)
        s.append([f"Members: {rs[0].get('node_members', '')}"])
        s["A2"].alignment = _WRAP
        s.append([c[1] for c in _COLS])
        for c in range(1, len(_COLS) + 1):
            cell = s.cell(row=3, column=c)
            cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT
        for r in rs:
            s.append([r.get(k, "") for k, _, _ in _COLS])
            status = r.get("status", "proposed")
            for c in range(1, len(_COLS) + 1):
                cell = s.cell(row=s.max_row, column=c)
                cell.alignment = _WRAP
                if status in _STATUS_FILL:
                    cell.fill = _STATUS_FILL[status]
        for i, (_, _, w) in enumerate(_COLS, 1):
            s.column_dimensions[get_column_letter(i)].width = w
        s.freeze_panes = "A4"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Vision excerpt export
# ---------------------------------------------------------------------------

_VISION_STATUS = {
    "verified":       ("bekreftet i uttrekket",        PatternFill("solid", fgColor="E2EFDA")),
    "verified_loose": ("bekreftet ((type, nr)-match)", PatternFill("solid", fgColor="DDEBF7")),
    "new_candidate":  ("NY KANDIDAT - sjekk tegning",  PatternFill("solid", fgColor="FFF2CC")),
    "suspect":        ("ukjent format / mulig hallusinasjon",
                       PatternFill("solid", fgColor="F8CBAD")),
}


def write_vision_xlsx(excerpt: dict, path: Path,
                      title: str = "Vision HAZOP excerpt") -> Path:
    """Export a verified vision excerpt (from ai.hazop_vision) to Excel:
    one sheet with the observations, one with per-tag verification, both
    colour-coded by verification status so the trust level travels with
    the file into the meeting."""
    wb = Workbook()

    # ---- observations ------------------------------------------------------
    ws = wb.active
    ws.title = "Observations"
    ws.append([title]); ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Model reading: {excerpt.get('summary', '')}"])
    ws["A2"].alignment = _WRAP
    ws.append([])
    head = ["#", "Deviation", "Observation", "Tags (status in next sheet)"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=4, column=c)
        cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT
    for i, obs in enumerate(excerpt.get("observations", []), 1):
        tags = ", ".join(t["tag"] for t in obs.get("tags", []))
        ws.append([i, obs.get("deviation", ""), obs.get("observation", ""), tags])
        for c in range(1, 5):
            ws.cell(row=ws.max_row, column=c).alignment = _WRAP
    for col, w in zip("ABCD", (4, 16, 70, 40)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    ws.append([])
    ws.append(["Vision-generated proposals verified against the tag register "
               "- for HAZOP team review, not an authority."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)

    # ---- per-tag verification ---------------------------------------------
    tv = wb.create_sheet("Tag verification")
    head = ["Tag", "Status", "Context"]
    tv.append(head)
    for c in range(1, len(head) + 1):
        cell = tv.cell(row=1, column=c)
        cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT

    def _row(tag_entry, context):
        label, fill = _VISION_STATUS.get(tag_entry.get("status", ""),
                                         (tag_entry.get("status", "?"), None))
        tv.append([tag_entry.get("tag", ""), label, context])
        for c in range(1, 4):
            cell = tv.cell(row=tv.max_row, column=c)
            cell.alignment = _WRAP
            if fill:
                cell.fill = fill

    for i, obs in enumerate(excerpt.get("observations", []), 1):
        for t in obs.get("tags", []):
            _row(t, f"observation {i}: {obs.get('deviation', '')}")
    for so in excerpt.get("possible_symbol_only", []):
        _row(so, f"symbol-only candidate: {so.get('symbol', '')}")
    for col, w in zip("ABC", (18, 36, 44)):
        tv.column_dimensions[col].width = w
    tv.freeze_panes = "A2"
    c = excerpt.get("tag_totals", {})
    tv.append([])
    tv.append([f"Totals: {c.get('verified', 0)} verified · "
               f"{c.get('verified_loose', 0)} verified (normalised) · "
               f"{c.get('new_candidate', 0)} new candidates · "
               f"{c.get('suspect', 0)} unknown format"])
    tv.cell(row=tv.max_row, column=1).font = Font(italic=True, size=9)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path